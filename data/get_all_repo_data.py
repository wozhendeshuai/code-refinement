import os
from datetime import datetime

import requests
from openpyxl import Workbook, load_workbook

# --- 配置 ---
# API endpoint for OpenHarmony organization repositories
API_BASE_URL = "https://gitcode.com/api/v5"
ORG_NAME = "OpenHarmony"
REPOS_ENDPOINT = f"{API_BASE_URL}/orgs/{ORG_NAME}/repos"

# Headers for the request (add authentication if needed)
HEADERS = {
    'Accept': 'application/json',
    'PRIVATE-TOKEN': 'dgnvGsNMKJ2F7Fvm72TVHWyC'
}
PER_PAGE = 100  # Number of repos per page (check API limits)

# Output file name with timestamp
TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILENAME = f"openharmony_repos_streaming_{TIMESTAMP}.xlsx"


# --- 配置结束 ---

def create_or_load_workbook(filename):
    """Creates a new workbook or loads an existing one."""
    if os.path.exists(filename):
        print(f"文件 {filename} 已存在，将加载它。")
        wb = load_workbook(filename)
        ws = wb.active
        # Determine the next row to write to (after header and existing data)
        start_row = ws.max_row + 1
        return wb, ws, start_row
    else:
        print(f"创建新文件 {filename}。")
        wb = Workbook()
        ws = wb.active
        # Write the header row
        headers = ['序号', 'id', 'full_name', 'created_at', 'language',
                   'internal', 'fork', 'html_url', 'forks_count',
                   'stargazers_count', 'watchers_count', 'default_branch',
                   'open_issues_count',
                   'project_creator', 'status',
                   'private', 'public', 'description', 'repo_json']
        ws.append(headers)
        wb.save(filename)
        return wb, ws, 2  # Start writing data from row 2


def flatten_and_write_batch(repos_data, workbook, worksheet, start_index):
    """
    Flattens a batch of repo data and appends it to the worksheet.
    Returns the next start index.
    """
    current_index = start_index
    rows_to_append = []
    for repo in repos_data:
        # Handle potential missing 'description' or other fields
        description = repo.get('description', '')
        # Ensure description is a string
        if not isinstance(description, str):
            description = str(description) if description is not None else ''

        flattened_repo = [
            current_index,  # 序号
            repo.get('id', ''),
            repo.get('full_name', ''),
            repo.get('created_at', ''),
            repo.get('language', ''),
            repo.get('internal', ''),
            repo.get('fork', ''),
            repo.get('html_url', ''),
            repo.get('forks_count', 0),
            repo.get('stargazers_count', 0),
            repo.get('watchers_count', 0),
            repo.get('default_branch', ''),
            repo.get('open_issues_count', 0),
            repo.get('project_creator', ''),
            repo.get('status', ''),
            repo.get('private', ''),
            repo.get('public', ''),  # Assuming default public if not present
            description,
            str(repo)
        ]
        rows_to_append.append(flattened_repo)
        current_index += 1

    # Append all rows in the batch at once for efficiency
    for row_data in rows_to_append:
        worksheet.append(row_data)

    try:
        workbook.save(OUTPUT_FILENAME)
        print(f"已保存第 {start_index} 到 {current_index - 1} 条记录。")
    except PermissionError:
        print(f"错误：无法写入文件 '{OUTPUT_FILENAME}'。请确保文件未被其他程序打开。")
        raise  # Re-raise to stop execution

    return current_index


def main():
    """Main function to orchestrate fetching and streaming data to Excel."""
    wb, ws, next_row_index = create_or_load_workbook(OUTPUT_FILENAME)

    page = 1
    total_fetched = 0
    print(f"开始获取 {ORG_NAME} 组织下的仓库列表，并实时写入 {OUTPUT_FILENAME}...")

    try:
        while True:
            params = {'page': page, 'per_page': PER_PAGE}
            print(f"正在获取第 {page} 页...")
            response = requests.get(REPOS_ENDPOINT, headers=HEADERS, params=params)
            response.raise_for_status()

            repos_data = response.json()

            # Check if the response is a list and not empty
            if isinstance(repos_data, list):
                if not repos_data:  # Empty list means no more data
                    print(f"第 {page} 页无数据，获取完成。")
                    break

                batch_size = len(repos_data)
                total_fetched += batch_size
                print(f"获取到第 {page} 页数据 ({batch_size} 个仓库)，正在处理并写入...")

                # Process and write this batch immediately
                next_row_index = flatten_and_write_batch(repos_data, wb, ws, next_row_index)

                print(f"第 {page} 页处理并写入完成。")
                page += 1
            else:
                # Likely an error object
                print(f"API 返回错误或非预期格式 (页 {page}): {repos_data}")
                break

    except requests.exceptions.RequestException as e:
        print(f"请求过程中发生网络错误: {e}")
    except Exception as e:
        print(f"处理数据或写入文件时发生错误: {e}")
    finally:
        # Ensure the workbook is saved upon exit (even if an error occurred)
        try:
            wb.save(OUTPUT_FILENAME)
            print(f"\n最终保存。总共尝试获取了 {total_fetched} 个仓库的信息。")
            print(f"数据已实时流式写入至 '{OUTPUT_FILENAME}'。")
        except:
            pass  # Ignore errors during final save if already handled


if __name__ == "__main__":
    main()
